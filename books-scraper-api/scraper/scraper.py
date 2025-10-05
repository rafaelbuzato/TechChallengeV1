"""
Web Scraper para Books to Scrape - Versão Otimizada
====================================================
Extrai informações de livros e salva em Excel/CSV

Uso:
    python scraper/scraper.py
"""

import requests
from bs4 import BeautifulSoup
import csv
import time
from typing import List, Dict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime


class BooksScraperSimple:
    """Scraper simplificado para Books to Scrape"""
    
    def __init__(self, base_url: str = "https://books.toscrape.com/"):
        self.base_url = base_url
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
    
    def get_page(self, url: str) -> BeautifulSoup:
        """Faz requisição e retorna BeautifulSoup"""
        try:
            response = self.session.get(url, timeout=10)
            response.raise_for_status()
            return BeautifulSoup(response.content, 'html.parser')
        except Exception as e:
            print(f"❌ Erro ao acessar {url}: {e}")
            return None
    
    def extract_rating(self, rating_classes: list) -> int:
        """Converte classe CSS para número de estrelas"""
        ratings = {'One': 1, 'Two': 2, 'Three': 3, 'Four': 4, 'Five': 5}
        for word in rating_classes:
            if word in ratings:
                return ratings[word]
        return 0
    
    def get_category(self, book_url: str) -> str:
        """Extrai categoria do livro"""
        soup = self.get_page(book_url)
        if not soup:
            return "Unknown"
        
        breadcrumb = soup.find('ul', class_='breadcrumb')
        if breadcrumb:
            links = breadcrumb.find_all('a')
            if len(links) >= 3:
                return links[2].text.strip()
        
        return "Unknown"
    
    def scrape_page(self, page_url: str) -> List[Dict]:
        """Extrai dados de uma página de listagem"""
        soup = self.get_page(page_url)
        if not soup:
            return []
        
        books = []
        articles = soup.find_all('article', class_='product_pod')
        
        print(f"📚 Encontrados {len(articles)} livros na página")
        
        for article in articles:
            try:
                # Título
                title_tag = article.find('h3').find('a')
                titulo = title_tag.get('title', '')
                
                # URL completa do livro
                book_url = self.base_url + 'catalogue/' + title_tag.get('href', '').replace('../', '')
                
                # Preço
                price_tag = article.find('p', class_='price_color')
                preco = price_tag.text.strip() if price_tag else 'N/A'
                
                # Rating
                rating_tag = article.find('p', class_='star-rating')
                rating = self.extract_rating(rating_tag.get('class', [])) if rating_tag else 0
                
                # Disponibilidade
                avail_tag = article.find('p', class_='instock availability')
                disponibilidade = avail_tag.text.strip() if avail_tag else 'N/A'
                
                # Imagem
                img_tag = article.find('img')
                imagem = self.base_url + img_tag.get('src', '').replace('../', '') if img_tag else 'N/A'
                
                # Categoria (precisa acessar página do livro)
                print(f"  → Extraindo categoria de: {titulo[:50]}...")
                categoria = self.get_category(book_url)
                
                books.append({
                    'titulo': titulo,
                    'preco': preco,
                    'rating': rating,
                    'disponibilidade': disponibilidade,
                    'categoria': categoria,
                    'imagem': imagem
                })
                
                # Pausa para não sobrecarregar servidor
                time.sleep(0.5)
                
            except Exception as e:
                print(f"⚠️  Erro ao processar livro: {e}")
                continue
        
        return books
    
    def scrape_all_pages(self, max_pages: int = 3) -> List[Dict]:
        """Extrai dados de múltiplas páginas"""
        all_books = []
        
        print("🕷️  Iniciando scraping...")
        print(f"📖 Páginas a extrair: {max_pages}")
        print("="*60)
        
        for page_num in range(1, max_pages + 1):
            page_url = self.base_url + f"catalogue/page-{page_num}.html"
            
            print(f"\n📄 Processando página {page_num}/{max_pages}...")
            
            books = self.scrape_page(page_url)
            
            if not books:
                print("⚠️  Nenhum livro encontrado. Finalizando...")
                break
            
            all_books.extend(books)
            print(f"✅ Total extraído até agora: {len(all_books)} livros")
            
            # Pausa entre páginas
            if page_num < max_pages:
                time.sleep(1)
        
        print("\n" + "="*60)
        print(f"🎉 Scraping concluído! Total: {len(all_books)} livros")
        
        return all_books
    
    def save_to_csv(self, books: List[Dict], filename: str = "data/books_data.csv"):
        """Salva dados em CSV"""
        if not books:
            print("⚠️  Nenhum dado para salvar")
            return
        
        # Cria diretório se não existir
        Path(filename).parent.mkdir(parents=True, exist_ok=True)
        
        fieldnames = ['titulo', 'preco', 'rating', 'disponibilidade', 'categoria', 'imagem']
        
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(books)
            
            print(f"💾 CSV salvo: {filename}")
        except Exception as e:
            print(f"❌ Erro ao salvar CSV: {e}")
    
    def save_to_excel(self, books: List[Dict], filename: str = "data/books_data.xlsx"):
        """Salva dados em Excel com formatação"""
        if not books:
            print("⚠️  Nenhum dado para salvar")
            return
        
        # Cria diretório se não existir
        Path(filename).parent.mkdir(parents=True, exist_ok=True)
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Livros"
            
            # Cabeçalhos
            headers = ['Título', 'Preço', 'Rating', 'Disponibilidade', 'Categoria', 'Imagem']
            ws.append(headers)
            
            # Formata cabeçalho
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Adiciona dados
            for book in books:
                ws.append([
                    book.get('titulo', ''),
                    book.get('preco', ''),
                    book.get('rating', 0),
                    book.get('disponibilidade', ''),
                    book.get('categoria', ''),
                    book.get('imagem', '')
                ])
            
            # Ajusta largura das colunas
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 25
            ws.column_dimensions['F'].width = 60
            
            # Alinha células
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                row[0].alignment = Alignment(horizontal='left', vertical='center')
                row[1].alignment = Alignment(horizontal='right', vertical='center')
                row[2].alignment = Alignment(horizontal='center', vertical='center')
                row[3].alignment = Alignment(horizontal='center', vertical='center')
                row[4].alignment = Alignment(horizontal='left', vertical='center')
                row[5].alignment = Alignment(horizontal='left', vertical='center')
            
            wb.save(filename)
            print(f"📊 Excel salvo: {filename}")
            
        except Exception as e:
            print(f"❌ Erro ao salvar Excel: {e}")


def main():
    """Função principal"""
    print("\n" + "="*60)
    print("🕷️  BOOKS TO SCRAPE - WEB SCRAPER")
    print("="*60)
    print(f"⏰ Início: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    # Cria scraper
    scraper = BooksScraperSimple()
    
    # Extrai dados (altere max_pages conforme necessário)
    books = scraper.scrape_all_pages(max_pages=3)
    
    # Salva em ambos os formatos
    if books:
        scraper.save_to_csv(books)
        scraper.save_to_excel(books)
        
        print("\n" + "="*60)
        print("✅ PROCESSO CONCLUÍDO!")
        print(f"📊 Total de livros: {len(books)}")
        print(f"📁 Arquivos salvos em: data/")
        print("="*60)
    else:
        print("\n❌ Nenhum livro foi extraído!")
    
    print(f"⏰ Fim: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")


if __name__ == "__main__":
    main()